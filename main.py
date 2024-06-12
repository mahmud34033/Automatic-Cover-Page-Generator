from pathlib import Path
from tkinter import Tk, Label, Button, Frame, Entry, Canvas, Toplevel, OptionMenu, Radiobutton, StringVar, IntVar
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import openpyxl
import pandas as pd
from tkcalendar import Calendar
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

window = Tk()
window.geometry("1000x600+130+70")
window.title("Automatic Cover Page Generator")
window.iconbitmap("assets/icon.ico")
window.resizable(False, False)


file_path = 'store_data.xlsx'
if not Path(file_path).is_file():
    wb = openpyxl.Workbook()
    sheet = wb.active
    headers = ["Serial_No.", "Doc_Type", "Template", "Course_Code", "Course_Title", "Topic_Name", "Exp_No", "Exp_Name", "Project_Title", "Teacher_Name", "Teacher_Designation", "Teacher_Department", "Student_Name", "Student_ID", "Section", "Semester", "Student_Department", "Submission_Date"]
    sheet.append(headers)
    wb.save(file_path)
        

#------------------------------------------------------------------------------------------
def validate_entries():
    if (type_var.get() == "Assignment" and dropdown_var1.get() == "Select a template:")\
        or (type_var.get() == "Lab Report" and dropdown_var2.get() == "Select a template:")\
        or (type_var.get() == "Project Report (Individual)" and dropdown_var3.get() == "Select a template:")\
        or (type_var.get() == "Project Report (Group)" and dropdown_var4.get() == "Select a template:"):
        messagebox.showerror("Error", "Please select a template!")
    elif any(entry.get() == "" for entry in [ccode_ent, ctitle_ent, tname_ent, tdsgn_ent, tdept_ent]):
        messagebox.showerror("Error", "Please fill in all the fields!")
    elif (type_var.get() == "Assignment" and topic_ent.get()=="")\
        or (type_var.get() == "Lab Report" and eno_ent.get()=="")\
        or (type_var.get() == "Lab Report" and ename_ent.get()=="")\
        or (type_var.get() == "Project Report (Individual)" and protitle_ent.get()=="")\
        or (type_var.get() == "Project Report (Group)" and protitle_ent.get()==""):
        messagebox.showerror("Error", "Please fill in all the fields!")
    elif (type_var.get() != "Project Report (Group)")\
        and any(entry.get() == "" for entry in [sname_ent, sid_ent, ssec_ent, semester_ent, sdept_ent]):
        messagebox.showerror("Error", "Please fill in all the fields!")
    elif type_var.get() == "Project Report (Group)"\
        and any(entry.get() == "" for entry in [sname_pg1_ent, sid_pg1_ent, sname_pg2_ent, sid_pg2_ent]):
            messagebox.showerror("Error", "Please fill in all the fields!")
    elif type_var.get() == "Project Report (Group)" and group_var.get() >= 3\
        and (sname_pg3_ent.get()=="" or sid_pg3_ent.get()==""):
            messagebox.showerror("Error", "Please fill in all the fields!")
    elif type_var.get() == "Project Report (Group)" and group_var.get() == 4\
        and (sname_pg4_ent.get()=="" or sid_pg4_ent.get()==""):
            messagebox.showerror("Error", "Please fill in all the fields!")
    elif date_ent.get() == "DD/MM/YYYY":
        messagebox.showerror("Error", "Please select a date!")
    else:
        if generate_pdf() != "Not Save":
            messagebox.showinfo("Success", "PDF saved successfully!")
            write_to_excel()
            clear_entries()

def write_to_excel():        
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    if dropdown_var1.get() != templates[0]:
        drop_var = dropdown_var1.get()
    elif dropdown_var2.get() != templates[0]:
        drop_var = dropdown_var2.get()
    elif dropdown_var3.get() != templates[0]:
        drop_var = dropdown_var3.get()
    elif dropdown_var4.get() != templates[0]:
        drop_var = dropdown_var4.get()
        
    data = [sheet.max_row, type_var.get(), drop_var, ccode_ent.get(), ctitle_ent.get(), topic_ent.get(), eno_ent.get(), ename_ent.get(), protitle_ent.get(), tname_ent.get(), tdsgn_ent.get(), tdept_ent.get(), sname_ent.get(), sid_ent.get(), ssec_ent.get(), semester_ent.get(), sdept_ent.get(), date_ent.get()]
    sheet.append(data)
    wb.save(file_path)
        
def generate_pdf():
    if dropdown_var1.get() == "Template-1" or dropdown_var2.get() == "Template-1" or dropdown_var3.get() == "Template-1" or dropdown_var4.get() == "Template-1":
        return design_1()
        # design_1()
    elif dropdown_var1.get() == "Template-2" or dropdown_var2.get() == "Template-2" or dropdown_var3.get() == "Template-2" or dropdown_var4.get() == "Template-2":
        return design_2()
    elif dropdown_var1.get() == "Template-3" or dropdown_var2.get() == "Template-3" or dropdown_var3.get() == "Template-3" or dropdown_var4.get() == "Template-3":
        return design_3()

def format_sentence(sentence, limit):
    words = sentence.split()
    lines = []
    current_line = ''
    line_length = 0

    for word in words:
        if line_length + len(word) + 1 <= limit:
            current_line += word + ' '
            line_length += len(word) + 1
        else:
            lines.append(current_line.strip())
            current_line = word + ' '
            line_length = len(word) + 1

    if current_line:
        lines.append(current_line.strip())

    return lines


def design_1():    
    design_file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=type_var.get()+"_Coverpage_"+sid_ent.get())
    if not design_file_path:
        return "Not Save"

    c = canvas.Canvas(design_file_path)
    diulogo = 'assets/diu.png'
    bglogo = 'assets/d1_bglogo.jpg'
    c.drawImage(diulogo, 40, 665, 200*2.5, 53*2.5)
    c.drawImage(bglogo, 85, 210, 420, 420)
    c.setStrokeColor("black")
    c.setLineWidth(2)
    
    if type_var.get()=="Assignment":
        c.line(252,610,366,610) # for d1 Assignment underline
    elif type_var.get()=="Lab Report":
        c.line(253,610,359,610) # for d1 labreport underline
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.line(231,610,369,610) # for d1 project report
    
    c.setStrokeColor("black")
    c.setLineWidth(2)
    c.line(66,444,167,444)
    c.setStrokeColor("black")
    c.setLineWidth(2)
    c.line(66,314,167,314)
    c.setStrokeColor("black")
    c.setLineWidth(2)
    c.line(66,112,195,112)
    c.translate(inch, inch)
    c.setFont("Helvetica-Bold", 20)
    
    if type_var.get()=="Assignment":
        c.drawString(2.5*inch, 7.5*inch, "Assignment")
    elif type_var.get()=="Lab Report":
        c.drawString(2.5*inch, 7.5*inch, "Lab Report")
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawString(2.2*inch, 7.5*inch, "Project Report")
        
    c.setFont("Helvetica-Bold", 16)
    c.drawString(-.1*inch, 7*inch, "Course Code: " + ccode_ent.get())
    c.drawString(-.1*inch, 6.66*inch, "Course Title: " + ctitle_ent.get())
    nxt_y = 6.32
    space_x = 1.34
    char_limit = 36
    
    if type_var.get()=="Assignment":     # for d1 Assignment
        c.drawString(-.1*inch, 6.32*inch, "Topic Name: ")
        input_sentence = topic_ent.get()
    elif type_var.get()=="Lab Report":
        c.drawString(-.1*inch, 6.32*inch, "Experiment No: " + eno_ent.get())
        c.drawString(-.1*inch, 5.98*inch, "Experiment Name: ")
        input_sentence = ename_ent.get()
        nxt_y = 5.98
        space_x = 1.85
        char_limit = 33
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawString(-.1*inch, 6.32*inch, "Project Title: ")
        input_sentence = protitle_ent.get()
        
    formatted_lines = format_sentence(input_sentence, char_limit)

    for line in formatted_lines:
        c.drawString(space_x*inch, nxt_y*inch, line)
        nxt_y -= 0.34

    c.drawString(-.1*inch, 5.2*inch, "Submitted To: ")
    c.drawString(1.35*inch, 4.90*inch, "Teacher Name : " + tname_ent.get())
    c.drawString(1.35*inch, 4.56*inch, "Designation     : " + tdsgn_ent.get())
    c.drawString(1.35*inch, 4.22*inch, "Department     : " + tdept_ent.get())
    c.drawString(1.35*inch, 3.88*inch, "Daffodil International University")
    c.drawString(-.1*inch, 3.4*inch, "Submitted By: ")
    
    if type_var.get()!="Project Report (Group)":
        c.drawString(1.35*inch, 3.08*inch, "Name : " + sname_ent.get())
        c.drawString(1.35*inch, 2.76*inch, "ID : " + sid_ent.get())
        c.drawString(1.35*inch, 2.44*inch, "Section : " + ssec_ent.get())
        c.drawString(1.35*inch, 2.12*inch, "Semester : " + semester_ent.get())
        c.drawString(1.35*inch, 1.80*inch, "Department : " + sdept_ent.get())
        c.drawString(1.35*inch, 1.48*inch, "Daffodil International University")
    else:
        c.drawString(1.35*inch, 3.08*inch, "1. " + sname_pg1_ent.get() + " (" + sid_pg1_ent.get() + ")")
        c.drawString(1.35*inch, 2.76*inch, "2. " + sname_pg2_ent.get() + " (" + sid_pg2_ent.get() + ")")
        if group_var.get() >= 3:
            c.drawString(1.35*inch, 2.44*inch, "3. " + sname_pg3_ent.get() + " (" + sid_pg3_ent.get() + ")")
        if group_var.get() == 4:
            c.drawString(1.35*inch, 2.12*inch, "4. " + sname_pg4_ent.get() + " (" + sid_pg4_ent.get() + ")")
    
    c.drawString(-.1*inch, 0.6*inch, "Submission Date: " + date_ent.get())
    c.save()

def design_2():    
    design_file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=type_var.get()+"_Coverpage_"+sid_ent.get())
    if not design_file_path:
        return "Not Save"

    c = canvas.Canvas(design_file_path)
    diulogo = 'assets/diu.png'
    bglogo = 'assets/d1_bglogo.jpg'
    c.drawImage(diulogo, x=117, y=677, width=200*1.8, height=53*1.8)
    c.drawImage(bglogo, x=179, y=303, width=233, height=233)
    c.setStrokeColor("#4472C4")
    c.setLineWidth(.4)
    c.line(x1=25, y1=815, x2=570, y2=815)
    c.line(x1=26, y1=814, x2=569, y2=814)
    c.line(x1=25, y1=815, x2=25, y2=25)
    c.line(x1=26, y1=814, x2=26, y2=26)
    c.line(x1=25, y1=25, x2=570, y2=25)
    c.line(x1=26, y1=26, x2=569, y2=26)
    c.line(x1=570, y1=815, x2=570, y2=25)
    c.line(x1=569, y1=814, x2=569, y2=26)
    c.line(x1=75, y1=650, x2=520, y2=650)
    c.line(x1=76, y1=649, x2=519, y2=649)
    c.line(x1=75, y1=650, x2=75, y2=619)
    c.line(x1=76, y1=649, x2=76, y2=620)
    c.line(x1=76, y1=620, x2=519, y2=620)
    c.line(x1=75, y1=619, x2=520, y2=619)
    c.line(x1=520, y1=650, x2=520, y2=619)
    c.line(x1=519, y1=649, x2=519, y2=620)
    c.line(x1=75, y1=110, x2=520, y2=110)
    c.line(x1=76, y1=109, x2=519, y2=109)
    c.line(x1=75, y1=85, x2=520, y2=85)
    c.line(x1=76, y1=86, x2=519, y2=86)
    c.line(x1=75, y1=110, x2=75, y2=85)
    c.line(x1=76, y1=109, x2=76, y2=86)
    c.line(x1=520, y1=110, x2=520, y2=85)
    c.line(x1=519, y1=109, x2=519, y2=86)
    c.translate(inch, inch)
    c.setFont("Times-Bold", 20)
    c.setFillColor("#4472C4")

    if type_var.get()=="Assignment":
        c.drawString(2.43*inch, 7.75*inch, "Assignment")
    elif type_var.get()=="Lab Report":
        c.drawString(2.43*inch, 7.75*inch, "Lab Report")
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawString(2.24*inch, 7.75*inch, "Project Report")

    c.setFont("Times-Bold", 14)
    c.drawString(.2*inch, 7.1*inch, "Course Code: " + ccode_ent.get())
    c.drawString(.2*inch, 6.75*inch, "Course Title: " + ctitle_ent.get())
    c.setFillColor("#000000")
    nxt_y = 6.4
    space_x = 1.34
    char_limit = 55
    
    if type_var.get()=="Assignment":
        c.drawString(.2*inch, 6.4*inch, "Topic Name: ")
        input_sentence = topic_ent.get()
    elif type_var.get()=="Lab Report":
        c.drawString(.2*inch, 6.4*inch, "Experiment No: " + eno_ent.get())
        c.drawString(.2*inch, 6.05*inch, "Experiment Name: ")
        input_sentence = ename_ent.get()
        nxt_y = 6.05
        space_x = 1.85
        char_limit = 50
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawString(.2*inch, 6.4*inch, "Project Title: ")
        input_sentence = protitle_ent.get()
    
    formatted_lines = format_sentence(input_sentence, char_limit)
    
    for line in formatted_lines:
        c.drawString(space_x*inch, nxt_y*inch, line)
        nxt_y -= 0.35
    
    c.setFont("Times-Bold", 16)
    c.setFillColor("#4472C4")
    c.drawString(.2*inch, 5.3*inch, "Submitted to: ")
    c.setFont("Times-Bold", 14)
    c.setFillColor("#000000")
    c.drawString(.2*inch, 4.95*inch, "Name: " + tname_ent.get())
    c.drawString(.2*inch, 4.6*inch, "Designation: " + tdsgn_ent.get())
    c.drawString(.2*inch, 4.25*inch, "Department of " + tdept_ent.get())
    c.drawString(.2*inch, 3.9*inch, "Daffodil International University")
    c.setFont("Times-Bold", 16)
    c.setFillColor("#4472C4")
    c.drawString(3.1*inch, 3.15*inch, "Submitted by: ")
    c.setFont("Times-Bold", 14)
    c.setFillColor("#000000")
    
    if type_var.get()!="Project Report (Group)":
        c.drawString(3.1*inch, 2.8*inch, "Name: " + sname_ent.get())
        c.drawString(3.1*inch, 2.45*inch, "ID: " + sid_ent.get())
        c.drawString(3.1*inch, 2.1*inch, "Section: " + ssec_ent.get())
        c.drawString(3.1*inch, 1.75*inch, "Semester: " + semester_ent.get())
        c.drawString(3.1*inch, 1.4*inch, "Department of " + sdept_ent.get())
        c.drawString(3.1*inch, 1.05*inch, "Daffodil International University")
    else:
        c.drawString(3.1*inch, 2.8*inch, "1. " + sname_pg1_ent.get() + " (" + sid_pg1_ent.get() + ")")
        c.drawString(3.1*inch, 2.45*inch, "2. " + sname_pg2_ent.get() + " (" + sid_pg2_ent.get() + ")")
        if group_var.get() >= 3:
            c.drawString(3.1*inch, 2.1*inch, "3. " + sname_pg3_ent.get() + " (" + sid_pg3_ent.get() + ")")
        if group_var.get() == 4:
            c.drawString(3.1*inch, 1.75*inch, "4. " + sname_pg4_ent.get() + " (" + sid_pg4_ent.get() + ")")
    
    c.setFont("Times-Roman", 16)
    c.drawString(.15*inch, .29*inch, "Submission Date: " + date_ent.get())
    c.save()


def design_3():
    design_file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=type_var.get()+"_Coverpage_"+sid_ent.get())
    if not design_file_path:
        return "Not Save"

    c = canvas.Canvas(design_file_path)
    diu_spelling = 'assets/diu_spelling.png'
    diulogo_only = 'assets/diulogo_only.png'
    c.drawImage(diu_spelling, x=174, y=720, width=109*2.25, height=40*2.25)
    c.drawImage(diulogo_only, x=236.5, y=599, width=123.5, height=123.5)
    c.setStrokeColor("#4472C4")
    c.setLineWidth(.9)
    c.line(x1=46, y1=590, x2=549, y2=590) # up
    c.line(x1=46, y1=555, x2=549, y2=555) #down
    c.translate(inch, inch)
    c.setFont("Times-Bold", 22)
    c.setFillColor("#09509E")

    if type_var.get()=="Assignment":
        c.drawCentredString(3.13*inch, 6.87*inch, "Assignment")
    elif type_var.get()=="Lab Report":
        c.drawCentredString(3.13*inch, 6.87*inch, "Lab Report")
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawCentredString(3.13*inch, 6.87*inch, "Project Report")

    c.setFont("Times-Bold", 16)
    c.setFillColor("#000000")
    c.drawString(-.38*inch, 6.2*inch, "Course Code: " + ccode_ent.get())
    c.drawString(-.38*inch, 5.81*inch, "Course Title: " + ctitle_ent.get())
    nxt_y = 5.42
    space_x = 0.93 # for topic and project
    char_limit = 52
    
    if type_var.get()=="Assignment":
        c.drawString(-.38*inch, 5.42*inch, "Topic Name: ")
        input_sentence = topic_ent.get()
    elif type_var.get()=="Lab Report":
        c.drawString(-.38*inch, 5.42*inch, "Experiment No: " + eno_ent.get())
        c.drawString(-.38*inch, 5.03*inch, "Experiment Name: ")
        input_sentence = ename_ent.get()
        nxt_y = 5.03
        space_x = 1.5 # for experiment
        char_limit = 50
    elif type_var.get()=="Project Report (Individual)" or type_var.get()=="Project Report (Group)":
        c.drawString(-.38*inch, 5.42*inch, "Project Title: ")
        input_sentence = protitle_ent.get()
    
    formatted_lines = format_sentence(input_sentence, char_limit)
    
    for line in formatted_lines:
        c.drawString(space_x*inch, nxt_y*inch, line)
        nxt_y -= 0.39

    c.setStrokeColor("#000000")
    c.setLineWidth(.4)
    if type_var.get()=="Project Report (Group)" and group_var.get() == 4: c.rect(-25, 58, 502, 250)   # TABLE OUTSIDE BOX ONLY
    else: c.rect(-25, 116, 502, 200) 
    c.setFillColor("#09509E")  # Fill the rectangle with the specified color
    c.rect(-25, 286, 502, 30, fill=1) # TABLE HEADER BOX ONLY
    c.line(x1=226, y1=316, x2=226, y2=116)  # table middle line
    c.setFont("Times-Bold", 18)
    c.setFillColor("#FFFFFF")
    c.drawCentredString(1.4*inch, 4.1*inch, "Submitted By")
    c.setFont("Times-Bold", 16)
    c.setFillColor("#000000")
    
    if type_var.get()!="Project Report (Group)":
        c.drawString(-.27*inch, 3.69*inch, "Name: " + sname_ent.get())
        c.drawString(-.27*inch, 3.3*inch, "Student ID: " + sid_ent.get())
        c.drawString(-.27*inch, 2.91*inch, "Section: " + ssec_ent.get())
        c.drawString(-.27*inch, 2.52*inch, "Semester: " + semester_ent.get())
        c.drawString(-.27*inch, 2.13*inch, "Department: " + sdept_ent.get())
        c.drawString(-.27*inch, 1.74*inch, "Daffodil International University")
    else:
        c.drawString(-.27*inch, 3.69*inch, "1. " + sname_pg1_ent.get())
        c.drawString(-.07*inch, 3.3*inch, "(" + sid_pg1_ent.get() + ")")
        c.drawString(-.27*inch, 2.91*inch, "2. " + sname_pg2_ent.get())
        c.drawString(-.07*inch, 2.52*inch, "(" + sid_pg2_ent.get() + ")")
        if group_var.get() >= 3:
            c.drawString(-.27*inch, 2.13*inch, "3. " + sname_pg3_ent.get())
            c.drawString(-.07*inch, 1.74*inch, "(" + sid_pg3_ent.get() + ")")
        if group_var.get() == 4:
            c.drawString(-.27*inch, 1.35*inch, "4. " + sname_pg4_ent.get())
            c.drawString(-.07*inch, 0.96*inch, "(" + sid_pg4_ent.get() + ")")
            c.line(x1=226, y1=316, x2=226, y2=58) # table middle line
            
    
    c.setFont("Times-Bold", 18)
    c.setFillColor("#FFFFFF")
    c.drawCentredString(4.88*inch, 4.1*inch, "Submitted To")
    c.setFont("Times-Bold", 16)
    c.setFillColor("#000000")
    c.drawString(3.22*inch, 3.69*inch, "Name: " + tname_ent.get())
    c.drawString(3.22*inch, 3.3*inch, "Designation: " + tdsgn_ent.get())
    c.drawString(3.22*inch, 2.91*inch, "Department: " + tdept_ent.get())
    c.drawString(3.22*inch, 2.52*inch, "Daffodil International University")
    if type_var.get()=="Project Report (Group)" and group_var.get() == 4: c.drawString(-.38*inch, 0.18*inch, "Date of Submission: " + date_ent.get())
    else: c.drawString(-.38*inch, 0.96*inch, "Date of Submission: " + date_ent.get())
    c.save()


def clear_entries():
    for entry in entries:
        entry.delete(0, 'end')
    date_ent.insert(0, "DD/MM/YYYY")
    type_var.set("Assignment")
    sel_type()
    window.focus()
# ---------------------------------------------------------------------------------------


# >>>>>>>>>>>>> LEFT Side Bar >>>>>>>>>>>

def change_frame(pressed):
    if pressed == "left_newgen_btn1":
        left_newgen_btn1.config(bg='#1A5276')
        left_history_btn1.config(bg='#2471A3')
        left_about_btn1.config(bg='#2471A3')
        table_frame.forget()
        upp.forget()
        about_frame.forget()
        main_frame.pack()
        
    elif pressed == "left_history_btn1":
        left_newgen_btn1.config(bg='#2471A3')
        left_history_btn1.config(bg='#1A5276')
        left_about_btn1.config(bg='#2471A3')  
        display_excel()
        main_frame.forget()
        about_frame.forget()
        
    elif pressed == "left_about_btn1":
        left_newgen_btn1.config(bg='#2471A3')
        left_history_btn1.config(bg='#2471A3')
        left_about_btn1.config(bg='#1A5276')
        table_frame.forget()
        upp.forget()
        main_frame.forget()
        about_frame.pack(side='left')

rrr = Frame(window)
inside_rrr = Frame(rrr, bg='#2471A3', width=140, height=600)
inside_rrr.grid(row=0, column=0, sticky='nsew')
icon_image = Image.open('assets/icon.png')
icon_image = icon_image.resize((512//15, 512//15))
canva = Canvas(inside_rrr, bg='#2471A3', highlightthickness=0, width=40, height=40)
canva.place(x=3, y=15)
icon_photo = ImageTk.PhotoImage(icon_image)
canva.create_image(25, 20, image=icon_photo)
Label(inside_rrr, text="Automatic\nCover Page\nGenerator", bg='#2471A3', fg='white', font=('Segoe UI',10,'bold'), justify='left').place(x=50, y=7)
left_newgen_btn1 = Button(inside_rrr, text="New Generate", width=19, height=2, bg='#1A5276', fg='white', activeforeground='white', activebackground="#1A5276", relief='flat', command=lambda: change_frame("left_newgen_btn1") )
left_newgen_btn1.place(x=0, y=130)
left_history_btn1 = Button(inside_rrr, text="History", width=19, height=2, bg='#2471A3', fg='white',  activeforeground='white', activebackground="#1A5276", relief='flat', command=lambda: change_frame("left_history_btn1"))
left_history_btn1.place(x=0, y=170)
left_about_btn1 = Button(inside_rrr, text="About", width=19, height=2, bg='#2471A3', fg='white',  activeforeground='white', activebackground="#1A5276", relief='flat', command=lambda: change_frame("left_about_btn1"))
left_about_btn1.place(x=0, y=210)
exit_btn = Button(inside_rrr, text="Exit", width=19, height=2, bg='#2471A3', fg='white',  activeforeground='white', activebackground="#1A5276", relief='flat', command=lambda: window.quit())
exit_btn.place(x=0, y=250)
Label(inside_rrr, text="© Developed by\nTeam_ASR", bg='#2471A3').place(x=22, y=535)
rrr.pack(side='left')
# <<<<<<<<<<<<<< LEFT Side Bar End <<<<<<<<<<<<



################################## NEW GENERATE Page Start #####################################
main_frame = Frame(window, width=1000, height=600)
main_frame.pack(fill='both')
right_frame1 = Frame(main_frame, width=860, height=600)
right_frame1.grid(row=0, column=1, sticky='nsew')
heading_right_frame1 = Frame(right_frame1,width=860, height=37, bg='#1ABC9C')
h1_label = Label(heading_right_frame1, text="New Generate", fg='white', bg='#1ABC9C', font=('Segoe UI',11,'bold')).place(x=390, y=5)
heading_right_frame1.pack(fill='both')
bottom_right_frame1 = Frame(right_frame1, width=860, height=73)
gen_btn = Button(bottom_right_frame1, text="Generate", command=validate_entries).place(x=405, y=10)
reset = Button(bottom_right_frame1, text="Reset", command=clear_entries).place(x=750, y=10)
bottom_right_frame1.pack(fill='both', side='bottom')
left_right_frame1 = Frame(right_frame1, width=228, height=490)
radio_button_frame1 = Frame(left_right_frame1, width=228, height=140)
radio_button_frame1.grid(row=0, column=0)
lbl1 = Label(radio_button_frame1, text="Select document type: ")
lbl1.pack()

# Radio button start
def gg():
    if type_var.get() == "Assignment":    # Assignment only
        sub_assign.tkraise()
        stu_all.tkraise()
    elif type_var.get() == "Lab Report":  # Lab Report only
        sub_lab.tkraise()
        stu_all.tkraise()
    elif type_var.get() == "Project Report (Individual)":  # project report only
        sub_project.tkraise()
        stu_all.tkraise()
    elif type_var.get() == "Project Report (Group)":  # new
        sub_project.tkraise()
        stu_project.tkraise()
        group_var.set(2)
        sel_members()

def sel_type():
    gg()
    
    topic_ent.delete(0, 'end')
    eno_ent.delete(0, 'end')
    ename_ent.delete(0, 'end')
    protitle_ent.delete(0, 'end')
    dropdown_var1.set(templates[0])
    dropdown_var2.set(templates[0])
    dropdown_var3.set(templates[0])
    dropdown_var4.set(templates[0]) # New Added for Group Project
    blank_frame.tkraise()
    if type_var.get() == "Assignment":
        drop_assignment_frame1.tkraise()
    elif type_var.get() == "Lab Report":
        drop_lab_frame1.tkraise()
    elif type_var.get() == "Project Report (Individual)":
        drop_project_frame1.tkraise()
    elif type_var.get() == "Project Report (Group)": # New Added for Group Project
        drop_project_frame2.tkraise()
 
document_types = ["Assignment", "Lab Report", "Project Report (Individual)", "Project Report (Group)"]
type_var = StringVar()
type_var.set("Assignment")
for type_name in document_types:
    radio = Radiobutton(radio_button_frame1, text=type_name, variable=type_var, value=type_name, command=sel_type)
    radio.pack(padx=55, anchor='w')
# Radio button End    

# Dropdown OptionMenu start
drop_assignment_frame1 = Frame(left_right_frame1, width=228, height=350)
drop_assignment_frame1.grid(row=1, column=0)
dp101 = Frame(drop_assignment_frame1, width=228, height=40)
dp101.grid(row=0, column=0)
dp102 = Frame(drop_assignment_frame1, width=228, height=310)
dp102.grid(row=1, column=0)

drop_lab_frame1 = Frame(left_right_frame1, width=228, height=350)
drop_lab_frame1.grid(row=1, column=0)
dp201 = Frame(drop_lab_frame1, width=228, height=40)
dp201.grid(row=0, column=0)
dp202 = Frame(drop_lab_frame1, width=228, height=310)
dp202.grid(row=1, column=0)

drop_project_frame1 = Frame(left_right_frame1, width=228, height=350)
drop_project_frame1.grid(row=1, column=0)
dp301 = Frame(drop_project_frame1, width=228, height=40)
dp301.grid(row=0, column=0)
dp302 = Frame(drop_project_frame1, width=228, height=310)
dp302.grid(row=1, column=0)

# New Added for Group Project
drop_project_frame2 = Frame(left_right_frame1, width=228, height=350)
drop_project_frame2.grid(row=1, column=0)
dp401 = Frame(drop_project_frame2, width=228, height=40)
dp401.grid(row=0, column=0)
dp402 = Frame(drop_project_frame2, width=228, height=310)
dp402.grid(row=1, column=0)


def create_template_dropdown(parent_frame, options):
    dropdown_frame = Frame(parent_frame, width=228, height=40)
    dropdown_frame.grid(row=0, column=0)
    dropdown_var = StringVar()
    dropdown_var.set(options[0])
    dropdown_menu = OptionMenu(dropdown_frame, dropdown_var, *options)
    dropdown_menu.pack()
    return dropdown_var

templates = ["Select a template:", "Template-1", "Template-2", "Template-3"]
dropdown_var1 = create_template_dropdown(dp101, templates)
dropdown_var2 = create_template_dropdown(dp201, templates)
dropdown_var3 = create_template_dropdown(dp301, templates)
dropdown_var4 = create_template_dropdown(dp401, templates) # New Added for Group Project

def open_new_window(image_path):
    new_window = Toplevel(window)
    new_window.title("New Window")
    new_image = Image.open(image_path)
    new_image = new_image.resize((2483//6, 3508//6))
    new_photo = ImageTk.PhotoImage(new_image)
    image_label = Label(new_window, image=new_photo)
    image_label.image = new_photo
    image_label.pack()

def create_image_label(parent_frame, image_path):
    image = Image.open(image_path)
    image = image.resize((2483//13, 3508//13))
    photo = ImageTk.PhotoImage(image)
    image_label = Label(parent_frame, image=photo)
    image_label.image = photo
    return image_label

image_label1 = create_image_label(dp102, "assets/D1A.jpg")
image_label2 = create_image_label(dp102, "assets/D2A.jpg")
image_label3 = create_image_label(dp102, "assets/D3A.jpg")
image_label4 = create_image_label(dp202, "assets/D1LAB.jpg")
image_label5 = create_image_label(dp202, "assets/D2LAB.jpg")
image_label6 = create_image_label(dp202, "assets/D3LAB.jpg")
image_label7 = create_image_label(dp302, "assets/D1P.jpg")
image_label8 = create_image_label(dp302, "assets/D2P.jpg")
image_label9 = create_image_label(dp302, "assets/D3P.jpg")
image_label10 = create_image_label(dp402, "assets/D1PG.jpg")
image_label11 = create_image_label(dp402, "assets/D2PG.jpg")
image_label12 = create_image_label(dp402, "assets/D3PG.jpg")


def create_view_button(parent_frame, image_path):
    button = Button(parent_frame, text="View template", command=lambda: open_new_window(image_path))
    button.pack_forget()
    return button

button1 = create_view_button(dp102, "assets/D1A.jpg")
button2 = create_view_button(dp102, "assets/D2A.jpg")
button3 = create_view_button(dp102, "assets/D3A.jpg")
button4 = create_view_button(dp202, "assets/D1LAB.jpg")
button5 = create_view_button(dp202, "assets/D2LAB.jpg")
button6 = create_view_button(dp202, "assets/D3LAB.jpg")
button7 = create_view_button(dp302, "assets/D1P.jpg")
button8 = create_view_button(dp302, "assets/D2P.jpg")
button9 = create_view_button(dp302, "assets/D3P.jpg")
button10 = create_view_button(dp402, "assets/D1PG.jpg")
button11 = create_view_button(dp402, "assets/D2PG.jpg")
button12 = create_view_button(dp402, "assets/D3PG.jpg")

def toggle_visibility(var, widget, template):
  if var.get() == template:
    widget.pack()
  else:
    widget.pack_forget()

dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, button1, "Template-1"))
dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, button2, "Template-2"))
dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, button3, "Template-3"))
dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, image_label1, "Template-1"))
dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, image_label2, "Template-2"))
dropdown_var1.trace("w", lambda *args: toggle_visibility(dropdown_var1, image_label3, "Template-3"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, button4, "Template-1"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, button5, "Template-2"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, button6, "Template-3"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, image_label4, "Template-1"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, image_label5, "Template-2"))
dropdown_var2.trace("w", lambda *args: toggle_visibility(dropdown_var2, image_label6, "Template-3"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, button7, "Template-1"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, button8, "Template-2"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, button9, "Template-3"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, image_label7, "Template-1"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, image_label8, "Template-2"))
dropdown_var3.trace("w", lambda *args: toggle_visibility(dropdown_var3, image_label9, "Template-3"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, button10, "Template-1"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, button11, "Template-2"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, button12, "Template-3"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, image_label10, "Template-1"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, image_label11, "Template-2"))
dropdown_var4.trace("w", lambda *args: toggle_visibility(dropdown_var4, image_label12, "Template-3"))




blank_frame = Frame(left_right_frame1, width=228, height=350)
blank_frame.grid(row=1, column=0)
drop_assignment_frame1.tkraise()
left_right_frame1.pack(fill='both', side='left')
# Dropdown OptionMenu End

# All Entry Start
right_right_frame1 = Frame(right_frame1, width=632, height=490)
up_rr_frame1 = Frame(right_right_frame1, width=632, height=41)
Label(up_rr_frame1, text="Course Code:").place(x=20, y=17)
ccode_ent = Entry(up_rr_frame1, width=10)
ccode_ent.place(x=115, y=17.5)
Label(up_rr_frame1, text="Course Title:").place(x=200, y=17)
ctitle_ent = Entry(up_rr_frame1, width=41)
ctitle_ent.place(x=288, y=17.5)
up_rr_frame1.pack(fill='both', side='top')
doctype_rr_frame1 = Frame(right_right_frame1, width=632, height=104)
sub_assign = Frame(doctype_rr_frame1, width=632, height=104)
sub_assign.grid(row=0, column=0)
Label(sub_assign, text="Topic Name:").place(x=20, y=40)
topic_ent = Entry(sub_assign, width=70)
topic_ent.place(x=115, y=42)
sub_lab = Frame(doctype_rr_frame1, width=632, height=104)
sub_lab.grid(row=0, column=0)
Label(sub_lab, text="Experiment No:").place(x=20, y=20)
eno_ent = Entry(sub_lab, width=10)
eno_ent.place(x=115, y=21)
Label(sub_lab, text="Experiment Name:").place(x=20, y=50)
ename_ent = Entry(sub_lab, width=66)
ename_ent.place(x=139, y=51)
sub_project = Frame(doctype_rr_frame1, width=632, height=104)
sub_project.grid(row=0, column=0)
Label(sub_project, text="Project Title:").place(x=20, y=40)
protitle_ent = Entry(sub_project, width=70)
protitle_ent.place(x=115, y=42)
sub_assign.tkraise()
doctype_rr_frame1.pack(fill='both')
teacher_rr_frame1 = Frame(right_right_frame1, width=632, height=124)
Label(teacher_rr_frame1, text="Teacher Information", font=('Segoe UI',9,'underline')).place(x=20, y=3)
Label(teacher_rr_frame1, text="Name:").place(x=20, y=30)
tname_ent = Entry(teacher_rr_frame1, width=70)
tname_ent.place(x=115, y=31)
Label(teacher_rr_frame1, text="Designation:").place(x=20, y=60)
tdsgn_ent = Entry(teacher_rr_frame1, width=70)
tdsgn_ent.place(x=115, y=61)
Label(teacher_rr_frame1, text="Department:").place(x=20, y=90)
tdept_ent = Entry(teacher_rr_frame1, width=70)
tdept_ent.place(x=115, y=91)
teacher_rr_frame1.pack(fill='both')
student_rr_frame1 = Frame(right_right_frame1, width=632, height=188)
stu_all = Frame(student_rr_frame1, width=632, height=188)
stu_all.grid(row=0, column=0)
Label(stu_all, text="Student Information", font=('Segoe UI',9,'underline')).place(x=20, y=3)
Label(stu_all, text="Name:").place(x=20, y=30)
sname_ent = Entry(stu_all, width=70)
sname_ent.place(x=115, y=31)
Label(stu_all, text="Student ID:").place(x=20, y=60)
sid_ent = Entry(stu_all, width=70)
sid_ent.place(x=115, y=61)
Label(stu_all, text="Section:").place(x=20, y=90)
ssec_ent = Entry(stu_all, width=70)
ssec_ent.place(x=115, y=91)
Label(stu_all, text="Semester:").place(x=20, y=120)
semester_ent = Entry(stu_all, width=70)
semester_ent.place(x=115, y=120)
Label(stu_all, text="Department:").place(x=20, y=150)
sdept_ent = Entry(stu_all, width=70)
sdept_ent.place(x=115, y=150)
student_rr_frame1.pack(fill='both')
stu_project = Frame(student_rr_frame1, width=632, height=188, bg='red')
stu_project.grid(row=0, column=0)
sub_stu_project_two = Frame(stu_project, width=632, height=114)
sub_stu_project_two.grid(row=0, column=0)
Label(sub_stu_project_two, text="Student Information", font=('Segoe UI',9,'underline')).place(x=20, y=3)
Label(sub_stu_project_two, text="Number of group members:").place(x=20, y=30)

def sel_members():
    if group_var.get() == 2:
        sub_stu_project_blank1.tkraise()
        sub_stu_project_blank2.tkraise()
    elif group_var.get() == 3:
        sub_stu_project_three.tkraise()
        sub_stu_project_blank2.tkraise()
    elif group_var.get() == 4:
        sub_stu_project_three.tkraise()
        sub_stu_project_four.tkraise()

group_var = IntVar()
group_var.set(2)
Radiobutton(sub_stu_project_two, text="2", variable=group_var, value=2, command=sel_members).place(x=200, y=30)
Radiobutton(sub_stu_project_two, text="3", variable=group_var, value=3, command=sel_members).place(x=250, y=30)
Radiobutton(sub_stu_project_two, text="4", variable=group_var, value=4, command=sel_members).place(x=300, y=30)
Label(sub_stu_project_two, text="1. Name:").place(x=20, y=60)
sname_pg1_ent = Entry(sub_stu_project_two, width=40)
sname_pg1_ent.place(x=115, y=61)
Label(sub_stu_project_two, text="ID:").place(x=380, y=60)
sid_pg1_ent = Entry(sub_stu_project_two, width=20)
sid_pg1_ent.place(x=415, y=61)
Label(sub_stu_project_two, text="2. Name:").place(x=20, y=90)
sname_pg2_ent = Entry(sub_stu_project_two, width=40)
sname_pg2_ent.place(x=115, y=91)
Label(sub_stu_project_two, text="ID:").place(x=380, y=90)
sid_pg2_ent = Entry(sub_stu_project_two, width=20)
sid_pg2_ent.place(x=415, y=91)
sub_stu_project_three = Frame(stu_project, width=632, height=35)
sub_stu_project_three.grid(row=1, column=0)
Label(sub_stu_project_three, text="3. Name:").place(x=20, y=6)
sname_pg3_ent = Entry(sub_stu_project_three, width=40)
sname_pg3_ent.place(x=115, y=7)
Label(sub_stu_project_three, text="ID:").place(x=380, y=6)
sid_pg3_ent = Entry(sub_stu_project_three, width=20)
sid_pg3_ent.place(x=415, y=7)
sub_stu_project_blank1 = Frame(stu_project, width=632, height=35)
sub_stu_project_blank1.grid(row=1, column=0)
sub_stu_project_four = Frame(stu_project, width=632, height=39)
sub_stu_project_four.grid(row=2, column=0)
Label(sub_stu_project_four, text="4. Name:").place(x=20, y=1)
sname_pg4_ent = Entry(sub_stu_project_four, width=40)
sname_pg4_ent.place(x=115, y=2)
Label(sub_stu_project_four, text="ID:").place(x=380, y=1)
sid_pg4_ent = Entry(sub_stu_project_four, width=20)
sid_pg4_ent.place(x=415, y=2)
sub_stu_project_blank2 = Frame(stu_project, width=632, height=39)
sub_stu_project_blank2.grid(row=2, column=0)

stu_all.tkraise()


def pick_date(event):
    global cal, date_window
    date_window = Toplevel()
    date_window.grab_set()
    date_window.title("Choose Date")
    date_window.geometry('250x220+590+370')
    cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/y")
    cal.place(x=0, y=0)
    submit_btn = Button(date_window, text="Submit", command=grab_date)
    submit_btn.place(x=95, y=190)
    
def grab_date():
    date_ent.delete(0, 'end')
    date_ent.insert(0, cal.get_date())
    date_window.destroy()

submission_rr_frame1 = Frame(right_right_frame1, width=632, height=33)
Label(submission_rr_frame1, text="Date of Submission:").place(x=20, y=0)
date_ent = Entry(submission_rr_frame1, width=13)
date_ent.place(x=150, y=1)
date_ent.insert(0, "DD/MM/YYYY")
date_ent.bind("<1>", pick_date)
submission_rr_frame1.pack(fill='both')
right_right_frame1.pack(fill='both', side='right')
right_frame1.tkraise()

entries = [
    ccode_ent, ctitle_ent, topic_ent, eno_ent, ename_ent, protitle_ent,
    tname_ent, tdsgn_ent, tdept_ent, sname_ent, sid_ent, ssec_ent,
    semester_ent, sdept_ent, date_ent,
    sname_pg1_ent, sname_pg2_ent, sname_pg3_ent, sname_pg4_ent,
    sid_pg1_ent, sid_pg2_ent, sid_pg3_ent, sid_pg4_ent
]
########################################## NEW GENERATE Page End ###############################################



############# History Page Start ##########

def display_excel():
    try:
        excel_data = pd.read_excel(file_path).fillna("-")
        # excel_data = excel_data.sort_values('Serial_No.', ascending=False)
        display_table(excel_data)
    except Exception as e:
        messagebox.showerror("Error", str(e))
        
def display_table(data):
    global table_frame, upp, edit_entry, delete_entry
    
    upp = Frame(window)
    forhead = Frame(upp, bg="#1ABC9C", width=860, height=37)
    Label(forhead, text="History", fg='white', bg='#1ABC9C', font=('Segoe UI',11,'bold')).place(x=405, y=5)
    forhead.pack(fill='both')
    fordown = Frame(upp, width=860, height=80)
    Label(fordown, text="To edit and regenerate, enter serial number: ").place(x=40, y=32)
    edit_btn = Button(fordown, text=" Enter ", command=edit_data)
    edit_btn.place(x=374, y=32)
    edit_entry = Entry(fordown, width=15, justify='center')
    edit_entry.place(x=280, y=35)
    delete_entry = Entry(fordown, width=10, justify='center')
    delete_entry.place(x=550, y=35)
    delete_btn = Button(fordown, text="Delete a row", command=delete_row_from_excel)
    delete_btn.place(x=614, y=32)
    clear_btn = Button(fordown, text="Clear all data", font=('Segoe UI',9,'underline'), relief='flat', command=clear_history)
    clear_btn.place(x=740, y=32)
    fordown.pack(fill='both')
    upp.pack(side='top')
    
    # Table frame start
    table_frame = ttk.Frame(window)
    table_frame.pack(padx=20, pady=1, side='left')
    treescroll_Y = ttk.Scrollbar(table_frame, orient='vertical')
    treescroll_Y.pack(side="right", fill="y")
    treescroll_X = ttk.Scrollbar(table_frame, orient='horizontal')
    treescroll_X.pack(side="bottom", fill="x")
    tree = ttk.Treeview(table_frame, columns=list(data.columns), yscrollcommand=treescroll_Y.set, xscrollcommand=treescroll_X.set, show="headings", height=20)
    tree.pack()
    
    for col in data.columns:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="center")
        
    for index, row in data.iterrows():
        values = [int(cell) if isinstance(cell, float) and cell.is_integer() else cell for cell in row]
        tree.insert("", "end", values=values)
        
    treescroll_Y.config(command=tree.yview)
    treescroll_X.config(command=tree.xview)
    # Table frame end

def edit_data():
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        serial_number = int(edit_entry.get()) + 1
        if serial_number > 1 and serial_number <= sheet.max_row:
            row = sheet[serial_number]
            
            doctype_value = row[1].value
            type_var.set(doctype_value)
            blank_frame.tkraise()
            if doctype_value == "Assignment":
                drop_assignment_frame1.tkraise()
                sub_assign.tkraise()
            elif doctype_value == "Lab Report":
                drop_lab_frame1.tkraise()
                sub_lab.tkraise()
            elif doctype_value == "Project Report (Individual)":
                drop_project_frame1.tkraise()
                sub_project.tkraise()
            elif doctype_value == "Project Report (Group)": # NEW FOR GROUP PROJECT
                drop_project_frame2.tkraise()
                sub_project.tkraise()
                stu_project.tkraise()

            template_value = row[2].value
            if template_value == "Template-1":
                if doctype_value == "Assignment":
                    dropdown_var1.set(templates[1])
                elif doctype_value == "Lab Report":
                    dropdown_var2.set(templates[1])
                elif doctype_value == "Project Report (Individual)":
                    dropdown_var3.set(templates[1])
                elif doctype_value == "Project Report (Group)":
                    dropdown_var4.set(templates[1])
            elif template_value == "Template-2":
                if doctype_value == "Assignment":
                    dropdown_var1.set(templates[2])
                elif doctype_value == "Lab Report":
                    dropdown_var2.set(templates[2])
                elif doctype_value == "Project Report (Individual)":
                    dropdown_var3.set(templates[2])
                elif doctype_value == "Project Report (Group)":
                    dropdown_var4.set(templates[2])
            elif template_value == "Template-3":
                if doctype_value == "Assignment":
                    dropdown_var1.set(templates[3])
                elif doctype_value == "Lab Report":
                    dropdown_var2.set(templates[3])
                elif doctype_value == "Project Report (Individual)":
                    dropdown_var3.set(templates[3])
                elif doctype_value == "Project Report (Group)":
                    dropdown_var4.set(templates[3])
            
            for i in range(15):
                entries[i].delete(0, 'end')
                entries[i].insert(0, str(row[i + 3].value))

            left_newgen_btn1.config(bg='#1A5276')
            left_history_btn1.config(bg='#2471A3')
            left_about_btn1.config(bg='#2471A3')
            table_frame.forget()
            upp.forget()
            main_frame.pack()
        else:
            messagebox.showinfo("Error", "Serial number out of range.")
        
    except FileNotFoundError:
        print(f"File '{file_path}' not found.")

def delete_row_from_excel():
    serial_number = int(delete_entry.get())
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    if serial_number > 0 and serial_number < sheet.max_row:
        result = messagebox.askyesno("Confirmation", f"Are you sure you want to delete row {serial_number}?")
    
        if result:
            sheet.delete_rows(serial_number + 1, 1)
            for i in range(serial_number, sheet.max_row):
                sheet.cell(i + 1, 1).value = i
            wb.save(file_path)
            messagebox.showinfo("Success", f"Row {serial_number} has been deleted successfully!")
            upp.forget()
            table_frame.forget()
            display_excel()
    else:
        messagebox.showinfo("Error", "Invalid serial number")

def clear_history():
    result = messagebox.askyesno("Confirmation", "Are you sure you want to clear the history?")
    if result:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet.delete_rows(2, sheet.max_row)
        wb.save(file_path)
        messagebox.showinfo("Success", "History has been deleted successfully!")
        upp.forget()
        table_frame.forget()
        display_excel()

upp = Frame(window)
upp.pack_forget()
table_frame = ttk.Frame(window)
table_frame.pack_forget()
############# History Page End ##############
        
        
################################ ABOUT Page Start #################################
about_frame = Frame(window)
about_head = Frame(about_frame, bg="#1ABC9C", width=860, height=37)
about_head.grid(row=0, column=0)
Label(about_head, text="About", fg='white', bg='#1ABC9C', font=('Segoe UI',11,'bold')).place(x=405, y=5)
about_body = Frame(about_frame, width=860, height=563)
about_body.grid(row=1, column=0)
appinfo = Frame(about_body, width=860, height=200)
appinfo.grid(row=0, column=0)
Label(appinfo, text="AppInfo:", font=('Segoe UI',10,'bold', 'underline')).place(x=48, y=15)
lbl_appinfo = Label(appinfo, width=115, height=10, text="Automatic Cover Page Generator: Create Professional Academic Cover Pages with Ease!\n\n\nMeet the \"Automatic Cover Page Generator\"! It's your solution for swiftly creating polished PDF cover pages for all your academic submissions.\n\n[+] Simply input your course details, teacher info, student particulars, and due dates, and choose from three pre-designed templates.\n[+] No internet? No worries! Our app works offline, ensuring convenience and consistency every time you need it.\n\nSay farewell to manual errors and hello to professional, hassle-free document formatting!", justify='left')
lbl_appinfo.place(x=22, y=39)
devinfo = Frame(about_body, width=860, height=363)
devinfo.grid(row=1, column=0)
Label(devinfo, text="Developer Info:", font=('Segoe UI',10,'bold', 'underline')).place(x=48, y=15)
user1_image = Image.open('assets/img_abdullah.png')
user1_image = user1_image.resize((90, 90))
user1_photo = ImageTk.PhotoImage(user1_image)
canva = Canvas(devinfo, highlightthickness=0, width=550, height=110)
canva.place(x=150, y=45)
canva.create_image(65, 60, image=user1_photo)
lbl_user1 = Label(devinfo, text="Abdullah Al Mahmud\nID: 222-15-6115\n(Project Manager,\nBackend Developer)\nContact: 01906-430317")
lbl_user1.place(x=156, y=165)
user2_image = Image.open('assets/img_sanjidul.png')
user2_image = user2_image.resize((90, 90))
user2_photo = ImageTk.PhotoImage(user2_image)
canva.create_image(275, 60, image=user2_photo)
lbl_user2 = Label(devinfo, text="Sanjidul Hasan\nID: 222-15-6434\n(Frontend Developer)\nContact: 01640-785322")
lbl_user2.place(x=365, y=165)
user3_image = Image.open('assets/img_rakib.png')
user3_image = user3_image.resize((90, 90))
user3_photo = ImageTk.PhotoImage(user3_image)
canva.create_image(485, 60, image=user3_photo)
lbl_user3 = Label(devinfo, text="Md. Mhamudul Islam Rakib\nID: 222-15-6437\n(UI/UX Designer)\nContact: 01996-629687")
lbl_user3.place(x=560, y=165)
about_frame.pack_forget()
############################### ABOUT Page End ######################################

window.mainloop()